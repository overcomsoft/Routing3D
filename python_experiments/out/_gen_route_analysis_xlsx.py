r"""DDW_AI_DB 설계배관(TB_ROUTE_PATH) 전수 분석 → 엑셀(.xlsx)
================================================================================
[실행 명령어]  (프로젝트 루트에서)
  .\.venv\Scripts\python.exe python_experiments\out\_gen_route_analysis_xlsx.py
  # 출력: out\ddw_route_analysis.xlsx (경로 인자로 변경 가능)
  .\.venv\Scripts\python.exe python_experiments\out\_gen_route_analysis_xlsx.py docs\X.xlsx

[이 스크립트가 하는 일]
--------------------------------------------------------------------------------
DDW_AI_DB 의 모든 설계배관(TB_ROUTE_PATH, ~7천건)을 읽어, 장비·유틸리티그룹·유틸리티·
출발/종단 PoC(ID·크기·소유자·소유자종류)·위치·벤딩수·길이를 한 행씩 엑셀로 내보낸다.
PoC 의 '소유자 종류'(MODEL/DUCT/LATERAL 등)는 TB_POCINSTANCES 조인으로 보강한다
(종단이 Duct 가 아니라 Damper 등으로 보이는 사례 분석용 — 종단 owner 는 설계상 덕트 부속
[댐퍼·엘보·테이크오프 등]일 수 있음).

[시트]
  1) "전체 경로"     : route_path 1행 = 엑셀 1행(요청 컬럼 + 보강 컬럼).
  2) "종단소유자 분석": 유틸리티그룹 × 종단 소유자종류 교차표(개수). 'Duct vs Damper' 한눈에.

[컬럼]  장비/장비태그/공정/유틸리티그룹/유틸리티 ·
        시작PoC(ID·크기·해석크기·소유자종류·위치) ·
        종단PoC(ID·크기·해석크기·소유자명·소유자종류[원문 첫토큰]·POC소유자타입·유틸·위치) ·
        벤딩수/총길이/ROUTE_PATH_GUID.

[DB]  PGHOST 등 env 우선, 기본 localhost/5432/DDW_AI_DB/postgres/dinno.
================================================================================
"""

from __future__ import annotations

import os
import sys

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _conn():
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"),
        port=int(os.environ.get("PGPORT", 5432)),
        dbname=os.environ.get("PGDATABASE", "DDW_AI_DB"),
        user=os.environ.get("PGUSER", "postgres"),
        password=os.environ.get("PGPASSWORD", "dinno"),
        options="-c client_encoding=UTF8 -c lc_messages=C",
    )


# 엑셀 헤더(한글) — SQL SELECT 순서와 1:1.
HEADERS = [
    "공정(PROCESS)", "베이(BAY)", "장비(EQUIPMENT_NAME)", "장비태그(EQUIPMENT_TAG)",
    "유틸리티그룹(UTILITY_GROUP)", "유틸리티(SOURCE_UTILITY)",
    "시작PoC ID(SOURCE_GUID)", "시작PoC 크기(SOURCE_SIZE)", "시작PoC 크기(POC)", "시작PoC 소유자타입",
    "시작 X", "시작 Y", "시작 Z",
    "종단PoC 소유자(TARGET_OWNER_NAME)", "종단 소유자종류", "종단PoC 소유자타입(POC)",
    "종단 유틸리티(TARGET_UTILITY)",
    "종단PoC ID(TARGET_GUID)", "종단PoC 크기(TARGET_SIZE)", "종단PoC 크기(POC)",
    "종단 X", "종단 Y", "종단 Z",
    "벤딩수(BEND_COUNT)", "총길이mm(TOTAL_LENGTH)", "ROUTE_PATH_GUID",
]

SQL = """
SELECT
  rp."PROCESS_NAME", rp."BAY", rp."EQUIPMENT_NAME", rp."EQUIPMENT_TAG",
  rp."UTILITY_GROUP", rp."SOURCE_UTILITY",
  rp."SOURCE_GUID", rp."SOURCE_SIZE", sp."PIPESIZE_NM", sp."OWNER_INSTANCE_TYPE",
  rp."SOURCE_POSX", rp."SOURCE_POSY", rp."SOURCE_POSZ",
  rp."TARGET_OWNER_NAME", tp."OWNER_INSTANCE_TYPE", rp."TARGET_UTILITY",
  rp."TARGET_GUID", rp."TARGET_SIZE", tp."PIPESIZE_NM",
  rp."TARGET_POSX", rp."TARGET_POSY", rp."TARGET_POSZ",
  rp."BEND_COUNT", rp."TOTAL_LENGTH", rp."ROUTE_PATH_GUID"
FROM "TB_ROUTE_PATH" rp
LEFT JOIN "TB_POCINSTANCES" sp ON sp."INSTANCE_ID" = rp."SOURCE_GUID"
LEFT JOIN "TB_POCINSTANCES" tp ON tp."INSTANCE_ID" = rp."TARGET_GUID"
ORDER BY rp."PROCESS_NAME", rp."EQUIPMENT_NAME", rp."UTILITY_GROUP", rp."SOURCE_UTILITY"
"""


def _owner_kind(name) -> str:
    """종단 소유자명 첫 토큰(공백/하이픈 전) = 소유자 종류. 예 'Damper-FMPVC-100A-Duct'→'Damper'."""
    if not name:
        return ""
    s = str(name).strip()
    for sep in ("-", " ", "_"):
        if sep in s:
            return s.split(sep)[0]
    return s


def _round(v):
    try:
        return round(float(v))
    except (TypeError, ValueError):
        return v


def main(argv):
    out = argv[1] if len(argv) > 1 else os.path.join("out", "ddw_route_analysis.xlsx")
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)

    conn = _conn()
    try:
        cur = conn.cursor()
        cur.execute(SQL)
        rows = cur.fetchall()
    finally:
        conn.close()
    print(f"TB_ROUTE_PATH 행수: {len(rows):,}")

    wb = Workbook()

    # ── 시트 1: 전체 경로 ──
    ws = wb.active
    ws.title = "전체 경로"
    head_fill = PatternFill("solid", fgColor="385B85")
    head_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # 종단소유자 분석 누적: (utility_group, owner_kind) → count.
    pivot: dict[tuple[str, str], int] = {}
    kinds: set[str] = set()
    groups: set[str] = set()

    for r in rows:
        (proc, bay, eq, eqtag, ugrp, sutil,
         sguid, ssize, spoc, sowner,
         sx, sy, sz,
         town, towner_type, tutil,
         tguid, tsize, tpoc,
         tx, ty, tz,
         bend, tlen, guid) = r
        okind = _owner_kind(town)
        ws.append([
            proc, bay, eq, eqtag, ugrp, sutil,
            sguid, ssize, spoc, sowner,
            _round(sx), _round(sy), _round(sz),
            town, okind, towner_type, tutil,
            tguid, tsize, tpoc,
            _round(tx), _round(ty), _round(tz),
            bend, _round(tlen), guid,
        ])
        g = ugrp or "(없음)"
        pivot[(g, okind)] = pivot.get((g, okind), 0) + 1
        kinds.add(okind)
        groups.add(g)

    # 컬럼 폭(대략).
    widths = [12, 8, 26, 16, 18, 14, 38, 14, 12, 16, 9, 9, 9, 30, 14, 18, 16, 38, 14, 12, 9, 9, 9, 9, 14, 38]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── 시트 2: 종단소유자 분석(유틸리티그룹 × 소유자종류) ──
    ws2 = wb.create_sheet("종단소유자 분석")
    kind_list = sorted(kinds, key=lambda k: -sum(pivot.get((g, k), 0) for g in groups))
    ws2.cell(1, 1, "유틸리티그룹 \\ 종단 소유자종류").font = head_font
    ws2.cell(1, 1).fill = head_fill
    for c, k in enumerate(kind_list, 2):
        cell = ws2.cell(1, c, k or "(없음)")
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.cell(1, len(kind_list) + 2, "합계").font = head_font
    ws2.cell(1, len(kind_list) + 2).fill = head_fill
    for rr, g in enumerate(sorted(groups), 2):
        ws2.cell(rr, 1, g).font = Font(bold=True)
        total = 0
        for c, k in enumerate(kind_list, 2):
            n = pivot.get((g, k), 0)
            if n:
                ws2.cell(rr, c, n)
            total += n
        ws2.cell(rr, len(kind_list) + 2, total).font = Font(bold=True)
    ws2.freeze_panes = "B2"
    ws2.column_dimensions["A"].width = 22
    for c in range(2, len(kind_list) + 3):
        ws2.column_dimensions[get_column_letter(c)].width = 11

    wb.save(out)
    print(f"저장 완료: {os.path.abspath(out)}")
    print(f"  시트1 '전체 경로' {len(rows):,}행 · 시트2 '종단소유자 분석' {len(groups)}그룹 × {len(kind_list)}종류")
    return 0


if __name__ == "__main__":
    for st in (sys.stdout, sys.stderr):
        try:
            st.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError):
            pass
    raise SystemExit(main(sys.argv))
